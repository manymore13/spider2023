# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# useful for handling different item types with a single interface

from itemadapter import ItemAdapter
import openpyxl
from openpyxl.styles import Alignment, Font, Color, GradientFill, PatternFill


class ExcelPipeline:
    title = ('研报名称', '机构名称', '发布时间', '行业', '研报地址')

    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

    def close_spider(self, spider):
        self.wb.save('行业研报.xlsx')

    def process_item(self, item, spider):
        sheet_name = item['tableName']
        cur_sheet = None
        if sheet_name not in self.wb.sheetnames:
            self.create_sheet(sheet_name)
        else:
            cur_sheet = self.wb[sheet_name]
        cur_sheet.append((item['title'], item['orgSName'], item['publishDate'], item['industryName'], item['pdfUrl']))

        return item

    def create_sheet(self, sheet_name):
        cur_sheet = self.wb.create_sheet(sheet_name, 0)
        cur_sheet.append(self.title)
        cell_range = cur_sheet['A1':'E1']
        for cell in cell_range:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name="黑体", sz=14, b=True, i=True, color=Color(indexed=0))
            cell.fill = PatternFill("solid", fgColor="AFEEEE")
            cell.height = 20
            cell.width = 20
